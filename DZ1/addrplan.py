# DZ1

from ipaddress import IPv4Network
from glob import glob
from re import match
from openpyxl import Workbook, styles, worksheet

dir = glob("E:\работа\Python\p4ne_training\config_files\*.txt")

iplist = []

for i in dir:
    with open(i) as f:
        for l in f:
            m = match(r" ip address ((?:\d{1,3}\.){3}\d{1,3}) ((?:\d{1,3}\.){3}\d{1,3})", l)
            if m:
                iplist.append(IPv4Network("%s/%s" % (m.group(1), m.group(2)), strict=False))

iplist = list(set(iplist))


def keyfunc(addr):
    ip = int(addr.network_address)
    mask = int(addr.netmask)
    return ip * 2 ** 32 + mask


iplistsorted = sorted(iplist, key=keyfunc)

wb = Workbook()
ws = wb.active

ws.append(['Network', 'Mask'])

for i in iplistsorted:
    ws.append(['%s' % i.network_address, '%s' % i.netmask])

ws['A1'].font = styles.Font( bold =True)
ws['B1'].font = styles.Font( bold =True)

wb.save("addrplan.xlsx")

